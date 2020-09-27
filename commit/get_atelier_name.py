# -*- coding: utf-8 -*-
"""
Created on Wed Aug 12 11:06:45 2020

@author: SAMSUNG
"""

from selenium import webdriver
import time
from urllib import parse
import openpyxl
mapp =['진천군', '삼척시', '김천시', '남구', '함양군', '고령군', '태안군', '완도군', '옹진군', 
       '양구군', '보성군', '무주군', '영덕군', '원주시', '김포시', '동해시', '양주시', '사하구', 
       '서대문구', '고흥군', '영암군', '과천시', '중구', '창녕군', '담양군', '광명시', '장성군', 
       '신안군', '성주군', '영월군', '고창군', '장수군', '노원구', '양산시', '해남군', '계양구', '울진군', 
       '무안군', '성동구', '상주시', '강남구', '산청군', '의정부시', '부평구', '횡성군', '익산시', '정선군', 
       '단양군', '강릉시', '의성군', '밀양시', '장흥군', '태백시', '용산구', '함안군', '통영시', '광산구', 
       '수영구', '화성시', '영등포구', '부천시', '청송군', '하동군', '함평군', '달성군', '연제구', 
       '가평군', '군포시', '제천시', '금산군', '충주시', '목포시', '임실군', '시흥시', '합천군', '완주군', 
       '순천시', '예산군', '군산시', '경산시', '포천시', '철원군', '양천구', '양양군', '마포구', '북구', 
       '강북구', '서귀포시', '홍성군', '봉화군', '사상구', '광주시', '춘천시', '의왕시', '괴산군', '오산시', 
       '관악구', '의령군', '계룡시', '당진시', '기장군', '청양군', '옥천군', '속초시', '홍천군', '안동시', 
       '영주시', '광양시', '부산진구', '화순군', '평창군', '강진군', '논산시', '영광군', '해운대구', 
       '중랑구', '울주군', '보령시', '여수시', '안성시', '구례군', '영동군', '연수구', '인제군', '청도군', 
       '동대문구', '서구', '유성구', '경주시', '거창군', '나주시', '강서구', '여주시', '증평군', '남원시', 
       '곡성군', '하남시', '순창군', '구리시', '서초구', '영천시', '구미시', '김해시', '영도구', '사천시', 
       '은평구', '달서구', '영양군', '화천군', '송파구', '남양주시', '서산시', '군위군', '예천군', '금천구', 
       '김제시', '남동구', '동구', '아산시', '진안군', '강동구', '진주시', '동작구', '평택시', '이천시', 
       '부여군', '진도군', '제주시', '파주시', '종로구', '정읍시', '금정구', '연천군', '도봉구', '동두천시', 
       '서천군', '성북구', '양평군', '구로구', '동래구', '울릉군', '대덕구', '수성구', '고성군', '공주시', 
       '광진구', '부안군', '음성군', '남해군', '거제시', '칠곡군', '강화군', '보은군', '문경시']



options = webdriver.ChromeOptions()
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")
driver = webdriver.Chrome('chromedriver', chrome_options=options)

work_file = openpyxl.load_workbook('아뜰리에정보.xlsx')

#work_file.create_sheet('전국 아뜰리에')
work_file_sheet = work_file.get_sheet_by_name('전국 아뜰리에')
work_file_sheet['A1']='지역'
work_file_sheet['B1']='아뜰리에 이름'
work_file_sheet['C1']='아뜰리에 네이버 코드'

#work_file.create_sheet('아뜰리에 별 정보')
area_file_sheet = work_file.get_sheet_by_name('아뜰리에 별 정보')
area_file_sheet['A1']='지역'
area_file_sheet['B1']='아뜰리에 개수'



for area in mapp:
    ###
    print(area,'하는중!')
    count=0
    ###
    
    ###각 지역별로 아뜰리에 개수를 찾아내고, 크롤링 할 개수를 조정하는 코드
    if len(area) == 2:
        url = 'https://store.naver.com/restaurants/list?filterId=r09&menu=%EB%94%94%EC%A0%80%ED%8A%B8&page=1&query={}%20%EB%A7%9B%EC%A7%91&sessionid=Sf0HuScwUZocKRimXiyUCN3p&sortingOrder=reviewCount'.format(parse.quote(area))
    elif len(area) >=3:
        ii = area[:-1]
        url = 'https://store.naver.com/restaurants/list?filterId=r09&menu=%EB%94%94%EC%A0%80%ED%8A%B8&page=1&query={}%20%EB%A7%9B%EC%A7%91&sessionid=Sf0HuScwUZocKRimXiyUCN3p&sortingOrder=reviewCount'.format(parse.quote(ii))
    driver.get(url)
    try:
        ateliers = driver.find_elements_by_xpath('/html/body/div/div/div[2]/div[1]/div/div[2]/div/div[1]/em')[0].text
    except:
        driver.refresh()
        ateliers = driver.find_elements_by_xpath('/html/body/div/div/div[2]/div[1]/div/div[2]/div/div[1]/em')[0].text
    atelier = int(ateliers.replace(',',''))
    if atelier >= 60:
        page_num = 4 #until 3page
        last_atelier = 21
    else:
        page_num = (atelier//20)+1 #all_page
        last_atelier=(atelier%20)+1
    ###
    
    ###
    for number in range(1, page_num):
        time.sleep(2)
        print(number,'page')
        if len(area) == 2:
            url = 'https://store.naver.com/restaurants/list?filterId=r09&menu=%EB%94%94%EC%A0%80%ED%8A%B8&page={}&query={}%20%EB%A7%9B%EC%A7%91&sessionid=Sf0HuScwUZocKRimXiyUCN3p&sortingOrder=reviewCount'.format(number, parse.quote(area))
        if len(area) >=3:
            ii = area[:-1]
            url = 'https://store.naver.com/restaurants/list?filterId=r09&menu=%EB%94%94%EC%A0%80%ED%8A%B8&page={}&query={}%20%EB%A7%9B%EC%A7%91&sessionid=Sf0HuScwUZocKRimXiyUCN3p&sortingOrder=reviewCount'.format(number, parse.quote(ii))
        driver.get(url)
        
        # 하나의 페이지 내에 아뜰리에의 개수를 알아내어 조정하는 코드
        if number == page_num:
            atelier_num = last_atelier
        else:
            atelier_num = 21
        last_len = work_file_sheet.max_row
        ###
        while True:
            try:
                for aa in range(1, atelier_num):
                    ###
                    row_len = work_file_sheet.max_row
                    ###
                
                    aa = driver.find_elements_by_xpath('/html/body/div/div/div[2]/div[1]/div/div[2]/ul/li[{}]/div/div/div[1]/span/a'.format(aa))
                    name=aa[0].get_attribute('title')
                    code=aa[0].get_attribute('href').split('&')[1].split('=')[-1]
                    print(name, code)
                
                    ###
                    work_file_sheet['A{}'.format(row_len+1)]=area
                    work_file_sheet['B{}'.format(row_len+1)]=name
                    work_file_sheet['C{}'.format(row_len+1)]=code
                    count+=1
                ###
            except:
                driver.refresh()
            if work_file_sheet.max_row > last_len:
                break
            else:
                break
        ### 
    area_row_len = area_file_sheet.max_row
    area_file_sheet['A{}'.format(area_row_len+1)]=area
    area_file_sheet['B{}'.format(area_row_len+1)]=count
    work_file.save('아뜰리에정보.xlsx')
#driver.quit()
    

            

        
        
        




