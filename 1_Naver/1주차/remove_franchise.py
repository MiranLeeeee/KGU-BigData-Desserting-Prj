# -*- coding: utf-8 -*-
"""
Created on Thu Aug 13 10:30:46 2020

@author: SAMSUNG
"""

import openpyxl
import collections
work_file = openpyxl.load_workbook('atelier_info.xlsx')
work_file_sheet = work_file.get_sheet_by_name('전국 아뜰리에')


row_len=work_file_sheet.max_row
franchise_list=[]
for atelier in range(2, row_len+1):
    aa = work_file_sheet['B{}'.format(atelier)].value.split()
    franchise_list.append(aa[0])
cc = collections.Counter(franchise_list)


franchise =[]
for atelier, value in cc.items():
    if value > 3:
        franchise.append(atelier)



w_file=openpyxl.load_workbook('remove_franchise_atelier_info.xlsx')
w_file.create_sheet('프렌차이즈 제거_전국 아뜰리에')
work_sheet = w_file.get_sheet_by_name('프렌차이즈 제거_전국 아뜰리에')
work_sheet['A1']='지역'
work_sheet['B1']='아뜰리에 이름'
work_sheet['C1']='아뜰리에 네이버 코드'

for atelier in range(2, row_len+1):
    aa = work_file_sheet['B{}'.format(atelier)].value
    bb = aa.split()[0]
    if bb not in franchise:
        work_len = work_sheet.max_row
        work_sheet['A{}'.format(work_len+1)]=work_file_sheet['A{}'.format(atelier)].value
        work_sheet['B{}'.format(work_len+1)]=aa
        work_sheet['C{}'.format(work_len+1)]=work_file_sheet['C{}'.format(atelier)].value
        print(aa)
    w_file.save('remove_franchise_atelier_info.xlsx')

w_file=openpyxl.load_workbook('remove_franchise_atelier_info.xlsx')
w_file.create_sheet('프렌차이즈 제거_지역 별 개수')
area_sheet = w_file.get_sheet_by_name('프렌차이즈 제거_지역 별 개수')
area_sheet['A1']='지역'
area_sheet['B1']='아뜰리에 개수'

work_sheet = w_file.get_sheet_by_name('프렌차이즈 제거_전국 아뜰리에')

row_len = work_sheet.max_row

area_list = []
for area in range(2,row_len+1):
    area_list.append(work_sheet['A{}'.format(area)].value)
area = collections.Counter(area_list)


for ar, va in area.items():
    r_len = area_sheet.max_row
    area_sheet['A{}'.format(r_len+1)]=ar
    area_sheet['B{}'.format(r_len+1)]=va

w_file.save('remove_franchise_atelier_info.xlsx')






