import time
import re
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl

'''
*Input
remove_franchise_atelier_info.xlsx: 각 아뜰리에 고유 아이디, 이름 데이터 파일

*Variables
storeId, storeName: 각 아뜰리에 고유 아이디(url key), 이름 -> storeIdList, storeList: 아이디, 이름 데이터 리스트 
reviewPage: 크롤링할 리뷰 페이지 수 
review(s): 각 아뜰리에 리뷰 데이터 -> reviewList: 리뷰 데이터 리스트
score(s): 각 아뜰리에 별점 데이터 -> scoreList: 별점 데이터 리스트
keyword(s), keywordString: 각 아뜰리에 테마키워드 데이터
description: 각 아뜰리에 요약정보(설명)
totalReviewPage: 각 아뜰리에 전체 리뷰 페이지 수

*Functions
getData: 각 아뜰리에 페이지별 리뷰데이터 크롤링
getKeywords: 각 아뜰리에 테마키워드 크롤링

*Output
atelier_reviews.xlsx: 각 아뜰리에 설명, 테마키워드, 리뷰, 평점 데이터 파일
'''

def getData(storeId, storeName, reviewPage):
    reviewList = []
    scoreList = []

    for i in range(0, reviewPage + 1):
        searchURL = "https://store.naver.com/restaurants/detail?entry=pll&id=" + storeId + "&query=" + storeName + "&tab=receiptReview&tabPage=" + str(i)
        driver.get(searchURL)

        reviews = driver.find_elements_by_class_name("review_txt")
        scores = driver.find_elements_by_class_name("score")

        for review in reviews:
            text = review.text
            #크롤링한 리뷰안에 이모지가 있을 경우 제거
            only_BMP_pattern = re.compile("["u"\U00010000-\U0010FFFF""]+", flags=re.UNICODE)
            text = only_BMP_pattern.sub(r'', text)
            reviewList.append(text)

        for score in scores:
            #실수값으로 저장
            scoreList.append(float(score.text))

    return reviewList, scoreList


def getKeywords(storeId, storeName):
    keywordString = ""
    description = ""

    searchURL = "https://store.naver.com/restaurants/detail?entry=pll&id=" + storeId + "&query=" + storeName + "&tab=main"
    driver.get(searchURL)
    keywords = driver.find_elements_by_class_name("kwd")
    description = driver.find_elements_by_class_name("ellipsis_area")[0]
    description = (description.find_elements_by_tag_name("span")[0]).text

    for keyword in keywords:
        keywordString += keyword.text

    return description, keywordString


storeList = []
storeIdList = []

load_wb = load_workbook("remove_franchise_atelier_info.xlsx", data_only=True)
load_ws = load_wb['프렌차이즈 제거_전국 아뜰리에']

for i in range(2, 6227):
    storeList.append(load_ws.cell(row=i,column=2).value)
    storeIdList.append(load_ws.cell(row=i,column=3).value)

path = "resource/chromedriver"
driver = webdriver.Chrome(path)
write_wb = openpyxl.Workbook()
write_wb.create_sheet('아뜰리에_리뷰평점')
write_wb.create_sheet('아뜰리에_이름설명키워드')

write_ws = write_wb.get_sheet_by_name('아뜰리에_리뷰평점')
write_ws['A1'] = '아뜰리에명'
write_ws['B1'] = '리뷰'
write_ws['C1'] = '평점'

write_info = write_wb.get_sheet_by_name('아뜰리에_이름설명키워드')
write_info['A1'] = '아뜰리에명'
write_info['B1'] = '설명'
write_info['C1'] = '키워드'

for i in range(0, 6225):
    name = ""
    reviewList = []
    scoreList = []
    description = ""
    keywords = ""
    searchURL = "https://store.naver.com/restaurants/detail?entry=pll&id=" + str(storeIdList[i]) + "&query=" + storeList[i] + "&tab=receiptReview"
    errorCheck = 0

    #크롤링 중 404 에러 발생 방지
    try:
        while True:
            try:
                driver.get(searchURL)
                errorCheck += 1
            except:
                driver.refresh()
            if errorCheck == 1:
                break
            else :
                break

        errorCheck = 0
        #페이지 전환 시 텀을 줌
        time.sleep(2)
        totalReviewPage = int(driver.find_elements_by_xpath("//*[@id='panel03']/div/div[2]/span")[0].text)
        reviewPage = 0

        #리뷰 개수가 30개~100개인 것만 크롤링
        if (totalReviewPage < 4):
            name = ""
            continue
        elif (totalReviewPage > 10):
            reviewPage = 3
        else:
            reviewPage = totalReviewPage

        while True:
            try:
                #각 아뜰리에 리뷰데이터와 평점 데이터 가져오기
                reviewList, scoreList = getData(str(storeIdList[i]), storeList[i], reviewPage)
                errorCheck += 1
            except:
                driver.refresh()
            if errorCheck == 1:
                break
            else :
                break

        errorCheck = 0
        while True:
            try:
                # 각 아뜰리에 요약정보(설명) 및 테마키워드 가져오기
                description, keywords = getKeywords(str(storeIdList[i]), storeList[i])
                errorCheck += 1
            except:
                driver.refresh()
            if errorCheck == 1:
                break
            else :
                break

        errorCheck = 0
        name = storeList[i]

        #가져온 각 데이터를 엑셀파일 셀에 저장
        for k in range(0, len(reviewList)):
            review = reviewList[k]
            score = scoreList[k]
            write_ws.append([name, review, score])

    except:
        driver.refresh()

    if len(name) > 1:
        write_info['A{}'.format(cnt+1)] = name
        write_info['B{}'.format(cnt+1)] = description
        write_info['C{}'.format(cnt+1)] = keywords

write_wb.save('atelier_reviews.xlsx')

