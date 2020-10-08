import time
import re
from openpyxl import load_workbook
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver

'''
*Input
atelier_info_reviews.xlsx: 리뷰가 30개 이상인 아뜰리에 고유 아이디, 이름 데이터 파일

*Variables
storeId, storeName (name): 각 아뜰리에 고유 아이디(url key), 이름 -> storeIdList, storeList: 아이디, 이름 데이터 리스트 
menu(s): 각 아뜰리에 메뉴 데이터 -> menuList: 메뉴 데이터 리스트
reviewPage: 크롤링할 리뷰 페이지 수 
title(s): 각 아뜰리에 리뷰 제목 데이터 -> titleList: 제목 데이터 리스트
review(s): 각 아뜰리에 리뷰 요약 데이터 -> reviewList: 리뷰 요약 데이터 리스트

*Functions
getMenu: 각 아뜰리에 메뉴 크롤링
getBlogReviews: 각 아뜰리에 블로그 리뷰(블로그 제목, 리뷰 요약) 크롤링  

*Output
atelier_menu.xlsx: 각 아뜰리에 메뉴, 블로그 리뷰 데이터 파일
'''


def getMenu(storeId, storeName):
    menuList = []
    searchURL = "https://store.naver.com/restaurants/detail?entry=plt&id=" + storeId + "&query=" + storeName + "&tab=menu"
    driver.get(searchURL)

    menus = driver.find_elements_by_class_name("tit")

    for menu in menus:
        text = menu.text
        if text != '':
            menuList.append(text)

    return menuList


def getBlogReviews(storeId, storeName, reviewPage):
    titleList = []
    reviewList = []

    for i in range(0, reviewPage + 1):
        searchURL = "https://store.naver.com/restaurants/detail?entry=pll&id=" + storeId + "&query=" + storeName + "&tab=fsasReview&tabPage=" + str(i)
        driver.get(searchURL)

        titles = driver.find_elements_by_tag_name(".tit>.name")
        reviews = driver.find_elements_by_class_name("ellp2")

        for title in titles:
            text = title.text
            titleList.append(text)

        for review in reviews:
            text = review.text
            only_BMP_pattern = re.compile("["u"\U00010000-\U0010FFFF""]+", flags=re.UNICODE)
            text = only_BMP_pattern.sub(r'', text)
            reviewList.append(text)

    return titleList, reviewList


storeList = []
storeIdList = []

load_wb = load_workbook("atelier_info_reviews.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

for i in range(1, 1511):
    storeList.append(load_ws.cell(row=i,column=1).value)
    storeIdList.append(load_ws.cell(row=i,column=2).value)

path = "resource/chromedriver"
driver = webdriver.Chrome(path)

write_wb = openpyxl.Workbook()
write_wb.create_sheet('아뜰리에_메뉴')
write_wb.create_sheet('아뜰리에_블로그리뷰')

write_menu = write_wb.get_sheet_by_name('아뜰리에_메뉴')
write_menu['A1'] = '아뜰리에명'
write_menu['B1'] = '메뉴'

write_blog = write_wb.get_sheet_by_name('아뜰리에_블로그리뷰')
write_blog['A1'] = '아뜰리에명'
write_blog['B1'] = '블로그제목'
write_blog['C1'] = '블로그리뷰'

titleList = []
reviewList = []
menuList = []

for i in range(0, len(storeList)):
    name = ""
    searchURL = "https://store.naver.com/restaurants/detail?entry=plt&id=" + str(storeIdList[i]) + "&query=" + storeList[i] + "&tab=fsasReview"
    errorCheck = 0

    try:
        while True:
            try:
                driver.get(searchURL)
                errorCheck += 1
            except:
                driver.refresh()
            if errorCheck == 1:
                break
            else:
                break

        errorCheck = 0
        time.sleep(2)
        totalReviewPage = int(driver.find_elements_by_xpath("//*[@id='panel04']/div/div/span")[0].text)
        reviewPage = 0

        #리뷰 30개 이하이면 다 크롤링
        if (totalReviewPage < 4):
            reviewPage = totalReviewPage
        #리뷰 31개 이상일 경우 30개만 크롤링
        elif (totalReviewPage >= 4):
            reviewPage = 3

        #아뜰리에 메뉴 크롤링
        while True:
            try:
                menuList = getMenu(str(storeIdList[i]), storeList[i])
                errorCheck += 1
            except:
                driver.refresh()
            if errorCheck == 1:
                break
            else:
                break

        errorCheck = 0
        #아뜰리에 블로그 리뷰(블로그 제목, 리뷰 요약) 크롤링
        while True:
            try:
                titleList, reviewList = getBlogReviews(str(storeIdList[i]), storeList[i], reviewPage)
                errorCheck += 1
            except:
                driver.refresh()
            if errorCheck == 1:
                break
            else:
                break

        errorCheck = 0
        name = storeList[i]
        for k in range(0, len(titleList)):
            title = titleList[k]
            review = reviewList[k]
            write_blog.append([name, title, review])

    except:
        driver.refresh()

    if len(name) > 1:
        for l in range(0, len(menuList)):
            menu = menuList[l]
            write_menu.append([name, menu])

    write_wb.save('atelier_menu.xlsx')